import openpyxl

def getRowCount(filename,sheetName):
    wb= openpyxl.load_workbook(filename)
    sh=wb[sheetName]
    return(sh.max_row)

def colCount(filename,sheetName):
    wb = openpyxl.load_workbook(filename)
    sh = wb[sheetName]
    return (sh.max_row)

def readData(filename,sheetName,rowNum,colNum):
    wb = openpyxl.load_workbook(filename)
    sh = wb[sheetName]
    return (sh.cell(row=rowNum,column=colNum).value)

def writeData(filename,sheetName,rowNum,colNum,data):
    wb = openpyxl.load_workbook(filename)
    sh = wb[sheetName]
    sh.cell(row=rowNum,column=colNum).value=data
    wb.save(filename)
